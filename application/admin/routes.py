from . import admin_bp
from io import BytesIO
# import io
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import logging
from flask import (
    abort,
    render_template,
    redirect,
    url_for,
    flash,
    request,
    Response,
    # make_response,
)
from flask_login import login_required, current_user
from ..models import Student, User, Subject, Result
from collections import defaultdict
from ..auth.forms import (
    EditStudentForm,
    ResultForm,
    SubjectForm,
    DeleteForm,
    SelectTermSessionForm,
    ApproveForm,
)
from ..helpers import (
    get_subjects_by_entry_class,
    update_results,
    calculate_results,
    db,
    random,
    string,
)
from datetime import datetime

# from weasyprint import HTML
from sqlalchemy.exc import SQLAlchemyError

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


@admin_bp.before_request
@login_required
def admin_before_request():
    if not current_user.is_admin:
        flash("You are not authorized to access this page.", "alert alert-danger")
        return redirect(url_for("main.index"))


@admin_bp.route("/dashboard")
def admin_dashboard():
    return render_template("admin/index.html")


@admin_bp.route("/admin/approve_students", methods=["GET", "POST"])
@login_required
def approve_students():
    if not current_user.is_admin:
        abort(403)  # Forbidden access

    approve_form = ApproveForm()
    deactivate_form = ApproveForm()
    regenerate_form = ApproveForm()

    students = Student.query.all()
    students = Student.query.all()
    students_by_class = defaultdict(list)
    for student in students:
        students_by_class[student.entry_class].append(student)

    return render_template(
        "admin/students/approve_students.html",
        students_by_class=students_by_class,
        approve_form=approve_form,
        deactivate_form=deactivate_form,
        regenerate_form=regenerate_form,
    )


@admin_bp.route("/admin/approve_student/<int:student_id>", methods=["POST"])
@login_required
def approve_student(student_id):
    if not current_user.is_admin:
        abort(403)  # Forbidden access

    form = ApproveForm()

    if form.validate_on_submit():
        student = Student.query.get_or_404(student_id)
        student.approved = True
        db.session.commit()
        flash(
            f"Student {student.first_name} {student.last_name} has been approved.",
            "alert alert-success",
        )
    else:
        flash("An error occurred. Please try again.", "alert alert-danger")
    return redirect(url_for("admins.approve_students"))


@admin_bp.route("/admin/deactivate_student/<int:student_id>", methods=["POST"])
@login_required
def deactivate_student(student_id):
    if not current_user.is_admin:
        abort(403)  # Forbidden access

    form = ApproveForm()
    if form.validate_on_submit():
        student = Student.query.get_or_404(student_id)
        student.approved = False
        db.session.commit()
        flash(
            f"Student {student.first_name} {student.last_name} has been deactivated.",
            "alert alert-success",
        )
    else:
        flash("An error occurred. Please try again.", "alert alert-danger")
    return redirect(url_for("admins.approve_students"))


@admin_bp.route("/admin/regenerate_password/<int:student_id>", methods=["POST"])
@login_required
def regenerate_password(student_id):
    if not current_user.is_admin:
        abort(403)  # Forbidden access

    form = ApproveForm()
    student = Student.query.get_or_404(student_id)
    if form.validate_on_submit() and student:

        # Generate a new temporary password
        new_temporary_password = "".join(
            random.choices(string.ascii_letters + string.digits, k=8)
        )
        # Update student's password
        student.password = new_temporary_password

        # Update user's password
        student.user.set_password(new_temporary_password)
        db.session.commit()
        flash(
            f"New password generated for {student.first_name} {student.last_name}: {new_temporary_password}",
            "alert alert-success",
        )
    else:
        flash("Student not found.", "alert alert-danger")
    return redirect(url_for("admins.approve_students"))


@admin_bp.route("/admin/students_by_class/<string:entry_class>")
@login_required
def students_by_class(entry_class):
    students = Student.query.filter_by(entry_class=entry_class).all()
    form = DeleteForm()  # Create an instance of the DeleteForm
    return render_template(
        "admin/classes/students_by_class.html",
        students=students,
        entry_class=entry_class,
        form=form,
    )


@admin_bp.route("/admin/manage_classes")
@login_required
def manage_classes():
    if not current_user.is_authenticated or not current_user.is_admin:
        return redirect(url_for("auth.login"))
    students = Student.query.all()
    return render_template("admin/classes/classes.html", students=students)


@admin_bp.route("/select_term_session/<int:student_id>", methods=["GET", "POST"])
@login_required
def select_term_session(student_id):
    student = Student.query.get_or_404(student_id)
    form = ResultForm()
    if form.validate_on_submit():
        term = form.term.data
        session = form.session.data
        return redirect(
            url_for("admins.manage_results", student_id=student.id, term=term, session=session)
        )
    return render_template("admin/results/select_term_session.html", form=form, student=student)


@admin_bp.route("/manage_results/<int:student_id>", methods=["GET", "POST"])
@login_required
def manage_results(student_id):
    if not current_user.is_authenticated or not current_user.is_admin:
        return redirect(url_for("auth.login"))

    try:
        student = Student.query.get_or_404(student_id)
        term = request.args.get("term")
        session = request.args.get("session")

        if not term or not session:
            return redirect(url_for("admins.select_term_session", student_id=student.id))

        form = ResultForm(term=term, session=session)
        subjects = get_subjects_by_entry_class(student.entry_class)

        if form.validate_on_submit():
            update_results(student, subjects, term, session, form)
            flash("Results updated successfully", "alert alert-success")
            return redirect(
                url_for(
                    "admins.manage_results", student_id=student.id, term=term, session=session
                )
            )

        results, grand_total, average, cumulative_average, results_dict = (
            calculate_results(student.id, term, session)
        )
        return render_template(
            "admin/results/manage_results.html",
            student=student,
            subjects=subjects,
            results=results,
            grand_total=grand_total,
            average=average,
            cumulative_average=cumulative_average,
            results_dict=results_dict,
            form=form,
            selected_term=term,
            selected_session=session,
        )
    except SQLAlchemyError as e:
        db.session.rollback()
        flash(f"Database error: {str(e)}", "alert alert-danger")
    except Exception as e:
        flash(f"An error occurred: {str(e)}", "alert alert-danger")
    return redirect(url_for("main.index"))


@admin_bp.route("/admin/delete_result/<int:result_id>", methods=["POST"])
@login_required
def delete_result(result_id):
    form = DeleteForm()
    if not current_user.is_authenticated or not current_user.is_admin:
        return redirect(url_for("auth.login"))

    result = Result.query.get_or_404(result_id)

    student_id = result.student_id
    term = result.term
    session = result.session

    try:
        db.session.delete(result)
        db.session.commit()
        flash("Result deleted successfully!", "alert alert-success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error deleting result: {e}", "alert alert-danger")

    return redirect(
        url_for(
            "admins.manage_results",
            form=form,
            student_id=student_id,
            term=term,
            session=session,
        )
    )


@admin_bp.route("/admin/manage_students", methods=["GET", "POST"])
@login_required
def manage_students():
    if not current_user.is_authenticated or not current_user.is_admin:
        return redirect(url_for("auth.login"))
    students = Student.query.all()
    return render_template("admin/students/student_admin.html", students=students)


@admin_bp.route("/admin/edit_student/<int:student_id>", methods=["GET", "POST"])
def edit_student(student_id):
    student = Student.query.get_or_404(student_id)
    form = EditStudentForm()

    if form.validate_on_submit():
        student.username = form.username.data
        student.entry_class = form.entry_class.data
        student.first_name = form.first_name.data
        student.last_name = form.last_name.data
        student.middle_name = form.middle_name.data
        student.gender = form.gender.data
        # Update other fields as needed

        # Update the username in the User model
        user = User.query.filter_by(id=student.user_id).first()
        user.username = form.username.data

        db.session.commit()
        flash("Student updated successfully!", "alert alert-success")
        return redirect(url_for("admins.students_by_class", entry_class=student.entry_class))
    elif request.method == "GET":
        form.username.data = student.username
        form.entry_class.data = student.entry_class
        form.first_name.data = student.first_name
        form.last_name.data = student.last_name
        form.middle_name.data = student.middle_name
        form.gender.data = student.gender
        # Populate other fields as necessary

    return render_template("admin/students/edit_student.html", form=form, student=student)


@admin_bp.route("/admin/delete_student/<int:student_id>", methods=["GET", "POST"])
def delete_student(student_id):
    form = DeleteForm()
    if form.validate_on_submit():
        if not current_user.is_authenticated or not current_user.is_admin:
            return redirect(url_for("auth.login"))

        student = Student.query.get_or_404(student_id)

        try:
            # Manually delete related results
            results = Result.query.filter_by(student_id=student.id).all()
            for result in results:
                db.session.delete(result)

            # Delete the associated user
            user = User.query.get(student.user_id)
            if user:
                db.session.delete(user)

            db.session.delete(student)
            db.session.commit()
            flash(
                "Student and associated results, along with user details, deleted successfully!",
                "alert alert-success",
            )
        except Exception as e:
            db.session.rollback()
            flash(f"Error deleting student: {e}", "alert alert-danger")

    return redirect(url_for("admins.students_by_class", entry_class=student.entry_class))


@admin_bp.route("/admin/manage_subjects", methods=["GET", "POST"])
@login_required
def manage_subjects():
    if not current_user.is_authenticated or not current_user.is_admin:
        return redirect(url_for("auth.login"))

    form = SubjectForm()
    if form.validate_on_submit():
        subjects_input = form.name.data
        subject_names = [name.strip() for name in subjects_input.split(",")]

        for subject_name in subject_names:
            for section in form.section.data:
                # Check if the subject already exists for the given section
                existing_subject = Subject.query.filter_by(
                    name=subject_name, section=section
                ).first()
                if existing_subject is None:
                    subject = Subject(name=subject_name, section=section)
                    db.session.add(subject)
        db.session.commit()
        flash("Subject(s) added successfully!", "alert alert-success")
        return redirect(url_for("admins.manage_subjects"))

    subjects = Subject.query.order_by(Subject.section).all()
    subjects_by_section = {}
    for subject in subjects:
        if subject.section not in subjects_by_section:
            subjects_by_section[subject.section] = []
        subjects_by_section[subject.section].append(subject)

    delete_form = DeleteForm()
    return render_template(
        "admin/subjects/subject_admin.html",
        form=form,
        subjects_by_section=subjects_by_section,
        delete_form=delete_form,
    )


@admin_bp.route("/admin/edit_subject/<int:subject_id>", methods=["GET", "POST"])
@login_required
def edit_subject(subject_id):
    subject = Subject.query.get_or_404(subject_id)
    form = SubjectForm(obj=subject)
    if form.validate_on_submit():
        subject.name = form.name.data
        subject.section = form.section.data
        db.session.commit()

        # Update all related results with the new subject name and section
        results = Result.query.filter_by(subject_id=subject.id).all()
        for result in results:
            result.subject_name = subject.name
            result.subject_section = subject.section
        db.session.commit()

        flash("Subject updated successfully!", "alert alert-success")
        return redirect(url_for("admins.manage_subjects"))

    return render_template("admin/subjects/edit_subject.html", form=form, subject=subject)


@admin_bp.route("/admin/delete_subject/<int:subject_id>", methods=["POST"])
def delete_subject(subject_id):
    if not current_user.is_authenticated or not current_user.is_admin:
        return redirect(url_for("auth.login"))

    form = DeleteForm()  # Instantiate the DeleteForm

    if form.validate_on_submit():
        try:
            # Find the subject
            subject = Subject.query.get_or_404(subject_id)

            # Delete all scores associated with the subject
            Result.query.filter_by(subject_id=subject_id).delete()

            # Delete the subject
            db.session.delete(subject)
            db.session.commit()

            flash(
                "Subject and associated scores deleted successfully!",
                "alert alert-success",
            )
        except Exception as e:
            db.session.rollback()
            flash(f"Error deleting subject: {e}", "alert alert-danger")

    return redirect(url_for("admins.manage_subjects"))


@admin_bp.route("/broadsheet/<string:entry_class>", methods=["GET", "POST"])
@login_required
def broadsheet(entry_class):
    if not current_user.is_authenticated or not current_user.is_admin:
        return redirect(url_for("auth.login"))

    form = SelectTermSessionForm()
    term = form.term.data if form.term.data else request.args.get("term")
    session = form.session.data if form.session.data else request.args.get("session")

    students = Student.query.filter_by(entry_class=entry_class).all()
    subjects = get_subjects_by_entry_class(entry_class=entry_class)

    broadsheet_data = []
    subject_averages = {subject.id: {"total": 0, "count": 0} for subject in subjects}

    for student in students:
        student_results = {
            "student": student,
            "results": {subject.id: {
                "class_assessment": "",
                "summative_test": "",
                "exam": "",
                "total": "",
                "grade": "",
                "remark": ""
            } for subject in subjects},
            "grand_total": "",
            "average": "",
            "position": "",
        }
        results = Result.query.filter_by(
            student_id=student.id, term=term, session=session
        ).all()

        grand_total = 0
        non_zero_subjects = 0
        for result in results:
            if result.subject_id in student_results["results"]:
                student_results["results"][result.subject_id] = {
                    "class_assessment": result.class_assessment if result.class_assessment is not None else "",
                    "summative_test": result.summative_test if result.summative_test is not None else "",
                    "exam": result.exam if result.exam is not None else "",
                    "total": result.total if result.total is not None and result.total > 0 else "",
                    "grade": result.grade if result.grade is not None else "",
                    "remark": result.remark if result.remark is not None else ""
                }
                if result.total > 0:
                    grand_total += result.total
                    non_zero_subjects += 1
                    subject_averages[result.subject_id]["total"] += result.total
                    subject_averages[result.subject_id]["count"] += 1
                student_results["position"] = result.position

        # Set grand_total and average with blank space if they are zero
        student_results["grand_total"] = grand_total if grand_total > 0 else ""
        average = grand_total / non_zero_subjects if non_zero_subjects > 0 else 0
        student_results["average"] = round(average, 1) if average > 0 else float('-inf')

        # student_results["average"] = round(average, 1) if average > 0 else ""

        # average = grand_total / non_zero_subjects if non_zero_subjects > 0 else 0
        # student_results["grand_total"] = grand_total
        # student_results["average"] = round(average, 1)
        broadsheet_data.append(student_results)

    # Calculate class averages for each subject
    for subject_id, values in subject_averages.items():
        values["average"] = (
            round(values["total"] / values["count"], 1) if values["count"] > 0 else 0
        )

    # Sort students by their average
    broadsheet_data.sort(key=lambda x: x["average"], reverse=True)

    return render_template(
        "admin/results/broadsheet.html",
        form=form,
        students=students,
        subjects=subjects,
        broadsheet_data=broadsheet_data,
        subject_averages=subject_averages,
        entry_class=entry_class,
    )

@admin_bp.route("/download_broadsheet/<string:entry_class>")
@login_required
def download_broadsheet(entry_class):
    if not current_user.is_authenticated or not current_user.is_admin:
        return redirect(url_for("auth.login"))

    term = request.args.get("term")
    session = request.args.get("session")

    students = Student.query.filter_by(entry_class=entry_class).all()
    subjects = get_subjects_by_entry_class(entry_class=entry_class)

    broadsheet_data = []
    subject_averages = {subject.id: {"total": 0, "count": 0} for subject in subjects}

    for student in students:
        student_results = {
            "student": student,
            "results": {subject.id: None for subject in subjects},
            "grand_total": 0,
            "average": 0,
            "position": None,
        }
        results = Result.query.filter_by(
            student_id=student.id, term=term, session=session
        ).all()

        grand_total = 0
        non_zero_subjects = 0
        for result in results:
            student_results["results"][result.subject_id] = result
            grand_total += result.total
            if result.total > 0:
                non_zero_subjects += 1
                subject_averages[result.subject_id]["total"] += result.total
                subject_averages[result.subject_id]["count"] += 1
            student_results["position"] = result.position

        average = grand_total / non_zero_subjects if non_zero_subjects > 0 else 0
        student_results["grand_total"] = grand_total
        student_results["average"] = round(average, 1)
        broadsheet_data.append(student_results)

    # Calculate class averages for each subject
    for subject_id, values in subject_averages.items():
        values["average"] = (
            round(values["total"] / values["count"], 1) if values["count"] > 0 else 0
        )

    # Sort students by their average in descending order
    broadsheet_data.sort(key=lambda x: x["average"], reverse=True)

    # Create Excel file
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"Broadsheet_{entry_class}_{term}"

    # Define styles
    header_font = Font(bold=True, size=14, name="Times New Roman")
    sub_header_font = Font(bold=True, size=12, name="Times New Roman")
    cell_font = Font(size=12, name="Times New Roman")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    alignment = Alignment(horizontal="left", vertical="center")

    # Add context information at the top
    sheet.append([f"Broadsheet for {entry_class} - Term: {term}, Session: {session}"])
    sheet.append([f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    sheet.append([])  # Blank row for separation

    # Write headers
    headers = ["Subjects"]
    for student_data in broadsheet_data:
        student = student_data["student"]
        headers.extend([f"{student.first_name} {student.last_name}", "", "", ""])
    headers.append("Class Average")
    sheet.append(headers)

    sub_headers = [""]
    for _ in broadsheet_data:
        sub_headers.extend(["C/A", "S/T", "Exam", "Total"])
    sub_headers.append("")
    sheet.append(sub_headers)

    for cell in sheet["1:1"]:
        cell.font = header_font
    for cell in sheet["2:2"]:
        cell.font = sub_header_font
    for cell in sheet["3:3"]:
        cell.font = cell_font
    sheet.merge_cells("A1:A2")

    # Write data
    for subject in subjects:
        row = [subject.name]
        for student_data in broadsheet_data:
            result = student_data["results"][subject.id]
            if result:
                row.extend(
                    [
                        result.class_assessment,
                        result.summative_test,
                        result.exam,
                        result.total,
                    ]
                )
            else:
                row.extend(["-", "-", "-", "-"])
        row.append(subject_averages[subject.id]["average"])
        sheet.append(row)

    # Write grand totals, averages, and positions
    sheet.append(
        ["Grand Total"]
        + sum(
            [
                ["", "", "", student_data["grand_total"]]
                for student_data in broadsheet_data
            ],
            [],
        )
        + [""]
    )
    sheet.append(
        ["Average"]
        + sum(
            [["", "", "", student_data["average"]] for student_data in broadsheet_data],
            [],
        )
        + [""]
    )
    sheet.append(
        ["Position"]
        + sum(
            [
                ["", "", "", student_data["position"]]
                for student_data in broadsheet_data
            ],
            [],
        )
        + [""]
    )

    # Set page orientation to landscape
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE

    # Apply styles
    for row in sheet.iter_rows(
        min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column
    ):
        for cell in row:
            cell.border = border
            cell.alignment = alignment
            cell.font = cell_font

    # Adjust column widths
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[column].width = adjusted_width

    # Save to a bytes buffer
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment;filename=Broadsheet_{entry_class}_{term}_{session}.xlsx"
        },
)