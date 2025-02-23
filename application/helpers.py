import random, string, time, openpyxl
from . import db
from io import BytesIO
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from flask import request, abort, current_app as app
from .models import Student, Subject, Result, Classes, Teacher
from functools import wraps
from datetime import datetime
from application.auth.forms import SubjectResultForm
from collections import defaultdict

login_attempts = {}

def generate_employee_id(section):
    """Generate a unique employee ID based on section and current year."""
    section_code = {
        "Nursery": "NRY",
        "Primary": "PRI",
        "Secondary": "SEC",
    }.get(section, "OTH")  # Default to "OTH" if section not found

    year = datetime.now().year
    last_teacher = Teacher.query.filter(
        Teacher.employee_id.like(f"AAIS-{section_code}-{year}-%")
    ).order_by(Teacher.id.desc()).first()

    if last_teacher:
        last_number = int(last_teacher.employee_id.split("-")[-1])
        new_number = f"{last_number + 1:03d}"
    else:
        new_number = "001"

    return f"AAIS-{section_code}-{year}-{new_number}"


def rate_limit(limit, per):
    """
    Decorator function that limits the rate at which a function can be called.

    Args:
        limit (int): The maximum number of calls allowed within the specified time period.
        per (int): The time period (in seconds) within which the maximum number of calls is allowed.

    Returns:
        function: The decorated function.

    Raises:
        HTTPException: If the maximum number of calls has been exceeded, a 429 Too Many Requests error is raised.
    """

    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            ip = request.remote_addr
            now = time.time()
            attempts = login_attempts.get(ip, [])
            attempts = [timestamp for timestamp in attempts if now - timestamp < per]

            if len(attempts) >= limit:
                abort(429)  # Too Many Requests

            attempts.append(now)
            login_attempts[ip] = attempts
            return f(*args, **kwargs)

        return wrapped

    return decorator

def group_students_by_class(students):
    """
    Groups students by class name based on the hierarchy.
    """
    students_classes = defaultdict(list)
    for student, class_name, _ in students:
        students_classes[class_name].append(student)

    # Sort classes by their hierarchy
    sorted_classes = sorted(
        students_classes.items(),
        key=lambda item: Classes.query.filter_by(name=item[0]).first().hierarchy,
    )
    return dict(sorted_classes)

def calculate_grade(total):
    if total >= 95:
        return "A+"
    elif total >= 80:
        return "A"
    elif total >= 70:
        return "B+"
    elif total >= 65:
        return "B"
    elif total >= 60:
        return "C+"
    elif total >= 50:
        return "C"
    elif total >= 40:
        return "D"
    elif total >= 30:
        return "E"
    else:
        return "F"


def generate_remark(total):
    if total >= 95:
        return "Outstanding"
    elif total >= 80:
        return "Excellent"
    elif total >= 70:
        return "Very Good"
    elif total >= 65:
        return "Good"
    elif total >= 60:
        return "Credit"
    elif total >= 50:
        return "Credit"
    elif total >= 40:
        return "Poor"
    elif total >= 30:
        return "Very Poor"
    else:
        return "Failed"


def generate_principal_remark(average):
    if average >= 90:
        remarks = [
            "Exceptional performance! You are an inspiration to others.",
            "Outstanding results! Keep setting high standards.",
            "Excellent work! You’ve proven yourself as a top performer.",
        ]
    elif average >= 80:
        remarks = [
            "Very impressive performance! Aim for even higher success.",
            "Great job! Your efforts are commendable.",
            "Consistently good results. Keep striving for excellence."
        ]
    elif average >= 70:
        remarks = [
            "Good performance! With a little more effort, you’ll excel further.",
            "You’ve done well, but there’s room for improvement.",
            "Nice results! Strive for greater achievements."
        ]
    elif average >= 60:
        remarks = [
            "Fair performance. Keep pushing for better results.",
            "You’re on the right track. Focus on improving your weaker areas.",
            "A decent effort. Keep working hard to achieve more."
        ]
    elif average >= 50:
        remarks = [
            "An average performance. Aim for consistent improvement.",
            "Your results are satisfactory, but you can do much better.",
            "A fair effort. Focus and determination will lead to success."
        ]
    elif average >= 40:
        remarks = [
            "Below average. Significant improvement is required.",
            "Your performance is concerning. Seek guidance to improve.",
            "You need to focus more on your studies to succeed."
        ]
    elif average >= 30:
        remarks = [
            "Poor results. A lot more effort is needed.",
            "Your performance requires immediate attention and action.",
            "Seek extra help to address your challenges effectively."
        ]
    else:
        remarks = [
            "Unacceptable performance. Take your studies seriously.",
            "Drastic improvement is needed to progress.",
            "This performance is not satisfactory. Immediate action is required."
        ]
    import random
    return random.choice(remarks)


def generate_teacher_remark(average):
    if average >= 90:
        remarks = [
            "Outstanding performance. Keep it up!",
            "Your dedication to learning is inspiring. Well done!",
            "An exemplary performance! Your efforts are commendable."
        ]
    elif average >= 80:
        remarks = [
            "A good performance! Focus on consistent excellence.",
            "Your hard work is paying off. Keep aiming higher!",
            "You’ve done very well. Aim to push your limits further."
        ]
    elif average >= 70:
        remarks = [
            "Good result, but practice more for perfection.",
            "You’re doing well, but focus on weaker areas to improve further.",
            "A good effort! Strive to achieve even better results."
        ]
    elif average >= 60:
        remarks = [
            "Fair result. Put in more effort.",
            "You’re progressing, but focus on difficult topics.",
            "With a bit more effort, your performance will improve significantly."
        ]
    elif average >= 50:
        remarks = [
            "An average performance. Work on improving your fundamentals.",
            "You’ve done okay. Additional practice will help.",
            "Keep working on weaker subjects to build a strong foundation."
        ]
    elif average >= 40:
        remarks = [
            "Below average understanding. Concentrate on improving your basics.",
            "Your performance is concerning. Seek help to address challenges.",
            "Weak results. Focus on understanding the core concepts."
        ]
    elif average >= 30:
        remarks = [
            "Very poor performance. A focused study plan is necessary.",
            "You need additional support to overcome your difficulties.",
            "Your results are below expectations. Put in much more effort."
        ]
    else:
        remarks = [
            "Unacceptable. Urgent attention to studies is required.",
            "A very poor performance. Take your academics more seriously.",
            "This is concerning. Take your academics more seriously."
        ]
    import random
    return random.choice(remarks)

def get_subjects_by_class_name(class_name, include_deactivated=False):
    """Get subjects based on the student's class name from history.

    Args:
        class_name (str): The student's class name.
        include_deactivated (bool): Whether to include deactivated subjects (for past sessions).

    Returns:
        list: A list of subjects based on the student's class name in ascending order.
    """

    # Determine section based on class_name
    if "Nursery" in class_name:
        query = Subject.query.filter_by(section="Nursery")
    elif "Basic" in class_name:
        query = Subject.query.filter_by(section="Basic")
    elif "SSS" in class_name:
        query = Subject.query.filter_by(section="Senior Secondary")
    else:
        query = Subject.query.filter_by(section="Secondary")

    # Include or exclude deactivated subjects based on the session context
    if not include_deactivated:
        query = query.filter_by(deactivated=False)

    # Order the results by subject name in ascending order
    return query.order_by(Subject.name.asc()).all()


def get_last_term(current_term):
    terms = ["First Term", "Second Term", "Third Term"]
    try:
        index = terms.index(current_term)
        return terms[index - 1] if index > 0 else None
    except ValueError:
        return None

def calculate_average(results):
    """
    Calculate the average score and number of subjects based on unique results.
    """
    # Use a set to avoid duplicate subject totals
    unique_results = {result.subject_id: result for result in results}.values()
    non_zero_results = [result.total for result in unique_results if result.total not in [None, 0]]
    total_sum = sum(non_zero_results)
    non_zero_subjects = len(non_zero_results)

    return total_sum / non_zero_subjects if non_zero_subjects > 0 else 0, non_zero_subjects


def calculate_cumulative_average(yearly_results):
    """
    Calculate the cumulative average over an academic year.
    Divides the total score by the number of subjects with non-zero totals.
    """
    # Filter out None values and zeroes, then sum and count the non-zero totals
    non_zero_results = [result.total for result in yearly_results if result.total not in [None, 0]]
    total_sum = sum(non_zero_results)
    non_zero_subjects = len(non_zero_results)

    return total_sum / non_zero_subjects if non_zero_subjects > 0 else 0


# Helper function to populate form with existing results
def populate_form_with_results(form, subjects, results_dict):
    for subject in subjects:
        result = results_dict.get(subject.id)
        subject_form = SubjectResultForm(
            subject_id=subject.id,
            class_assessment=result.class_assessment if result else 0,
            summative_test=result.summative_test if result else 0,
            exam=result.exam if result else 0,
            total=result.total if result else 0,
            grade=result.grade if result else "",
            remark=result.remark if result else "",
        )
        form.subjects.append_entry(subject_form)

def update_results(student, term, session_year, form, result_form):
    """
    Update results for each subject, ensuring consistency for remarks and calculations.
    """
    next_term_begins = result_form.next_term_begins.data
    date_issued = result_form.date_issued.data
    position = result_form.position.data


    for subject_form in form.subjects:
        subject_id = subject_form.subject_id.data
        data = {
            "class_assessment": subject_form.class_assessment.data,
            "summative_test": subject_form.summative_test.data,
            "exam": subject_form.exam.data,
            "next_term_begins": next_term_begins,
            "date_issued": date_issued,
            "position": position,
        }
        save_result(student.id, subject_id, term, session_year, data)

def save_result(student_id, subject_id, term, session_year, data):
    """
    Save or update a result for a student, ensuring consistency.
    """
    try:
        class_assessment = data.get("class_assessment")
        summative_test = data.get("summative_test")
        exam = data.get("exam")
        next_term_begins = data.get("next_term_begins")
        position = data.get("position")
        date_issued = data.get("date_issued")

        class_assessment_value = None if not class_assessment else int(class_assessment)
        summative_test_value = None if not summative_test else int(summative_test)
        exam_value = None if not exam else int(exam)

        if all(value in [None, 0] for value in [class_assessment_value, summative_test_value, exam_value]):
            total = None
        else:
            total = (class_assessment_value or 0) + (summative_test_value or 0) + (exam_value or 0)

        grade = calculate_grade(total) if total is not None else ''
        remark = generate_remark(total) if total is not None else ''

        result = Result.query.filter_by(
            student_id=student_id, subject_id=subject_id, term=term, session=session_year
        ).first()

        if result:
            # Update existing result
            result.class_assessment = class_assessment_value
            result.summative_test = summative_test_value
            result.exam = exam_value
            result.total = total
            result.grade = grade
            result.remark = remark
            result.next_term_begins = next_term_begins
            result.position = position
            result.date_issued = date_issued
        else:
            # Create a new result record
            new_result = Result(
                student_id=student_id,
                subject_id=subject_id,
                term=term,
                session=session_year,
                class_assessment=class_assessment_value,
                summative_test=summative_test_value,
                exam=exam_value,
                total=total,
                grade=grade,
                remark=remark,
                next_term_begins=next_term_begins,
                position=position,
                date_issued=date_issued,
            )
            db.session.add(result)

        db.session.commit()
         # Now calculate averages and grand total after commit
        grand_total, average, cumulative_average, last_term_average, subjects_offered = calculate_results(student_id, term, session_year)

        # Update the results with the calculated averages and grand total
        for result in Result.query.filter_by(student_id=student_id, term=term, session=session_year):
            result.grand_total = grand_total
            result.term_average = average
            result.cumulative_average = cumulative_average
            result.last_term_average = last_term_average
            result.subjects_offered = subjects_offered

        # Commit the final updates
        db.session.commit()
        return result  # Return the saved/updated result for further use
    except Exception as e:
        app.logger.error(f"Error in save_result: {str(e)}")
        raise

def calculate_results(student_id, term, session_year):
    """
    Calculate student results and averages, and ensure remarks are stored once.
    """
    # Fetch results for the current term and session
    results = Result.query.filter_by(
        student_id=student_id, term=term, session=session_year
    ).all()

    # Ensure unique results per subject
    unique_results = {result.subject_id: result for result in results}.values()

    # Calculate grand total and average
    grand_total = sum(result.total for result in unique_results if result.total is not None)
    average, subjects_offered = calculate_average(unique_results) if unique_results else (0, 0)
    average = round(average, 1) if average else 0


    # Calculate cumulative average
    yearly_results = Result.query.filter_by(student_id=student_id, session=session_year).all()
    unique_yearly_results = {result.subject_id: result for result in yearly_results}.values()
    cumulative_average = round(calculate_cumulative_average(unique_yearly_results), 1)

    # Last term average
    last_term = get_last_term(term)
    last_term_results = Result.query.filter_by(
        student_id=student_id, term=last_term, session=session_year
    ).all()
    unique_last_term_results = {result.subject_id: result for result in last_term_results}.values()
    last_term_average = round(calculate_average(unique_last_term_results)[0], 1) if last_term_results else None

    # Remarks
    principal_remark = generate_principal_remark(average)
    teacher_remark = generate_teacher_remark(average)

    # Apply remarks to the first result record if not already present
    if results and not results[0].principal_remark:
        results[0].principal_remark = principal_remark
        results[0].teacher_remark = teacher_remark
        db.session.commit()

    return grand_total, average, cumulative_average, last_term_average, subjects_offered

def prepare_broadsheet_data(students, subjects, term, session_year):
    broadsheet_data = []
    subject_averages = {subject.id: {"total": 0, "count": 0} for subject in subjects}

    for student in students:
        student_results = {
            "student": student,
            "results": {subject.id: None for subject in subjects},
            "grand_total": 0,
            "average": 0,
            "position": None,
        }

        # Fetch all results for the student
        results = Result.query.filter_by(student_id=student.id, term=term, session=session_year).all()

        for result in results:
            if result.subject_id in student_results["results"]:
                student_results["results"][result.subject_id] = result

                if result.total is not None and result.total > 0:
                    subject_averages[result.subject_id]["total"] += result.total
                    subject_averages[result.subject_id]["count"] += 1

            # Update grand_total, average, and position with defaults if values are missing
            student_results["grand_total"] = (
                result.grand_total if result.grand_total not in (None, "") else 0
            )
            student_results["average"] = (
                result.term_average if result.term_average not in (None, "") else 0
            )
            student_results["position"] = result.position if result.position else None

        broadsheet_data.append(student_results)

    # Calculate averages for each subject
    for subject_id, values in subject_averages.items():
        values["average"] = (
            round(values["total"] / values["count"], 1) if values["count"] > 0 else 0
        )

    # Safely sort the broadsheet by student averages
    broadsheet_data.sort(
        key=lambda x: float(x["average"]) if isinstance(x["average"], (int, float)) else 0,
        reverse=True,
    )

    return broadsheet_data, subject_averages


def generate_excel_broadsheet(class_name, term, session_year, broadsheet_data, subjects, subject_averages):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"Broadsheet_{class_name}_{term}"[:31]

    # Define styles
    header_font = Font(bold=True, size=16, name="Times New Roman")
    sub_header_font = Font(bold=True, size=12, name="Times New Roman")
    cell_font = Font(bold=True, size=12, name="Times New Roman")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    alignment = Alignment(horizontal="center", vertical="center")

    # Headers, Sub-Headers, and Data

    # Merge cells from A1 to the column for Class Average
    last_column = (
        2 + len(broadsheet_data) * 4
    )
    sheet.merge_cells(
        start_row=1, start_column=2, end_row=1, end_column=last_column
    )
    sheet["B1"].value = (
        f"Broadsheet for {class_name} - Term: {term}, Academic Session: {session_year}  -  Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    )
    sheet["B1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["B1"].font = header_font

    # Set column widths
    subject_column_letter = get_column_letter(1)
    sheet.column_dimensions[subject_column_letter].width = (
        30  # Set subject column width to 30cm
    )

    class_average_column_letter = get_column_letter(last_column)
    sheet.column_dimensions[class_average_column_letter].width = (
        15  # Set class average width to 15 cm
    )

    # Write headers (now student names merged and centralized)
    headers = [""]
    for student_data in broadsheet_data:
        student = student_data["student"]
        headers.extend([f"{student.first_name} {student.last_name}", "", "", ""])
    headers.append("Class Average")
    sheet.append(headers)

    # Merge and center student names across their 4 corresponding columns
    for i, student_data in enumerate(broadsheet_data, start=2):
        start_col = 2 + (i - 2) * 4
        end_col = start_col + 3

        # Set the width for C/A (start_col), S/T (start_col + 1), Exam (start_col + 2), and Total (start_col + 3)
        sheet.column_dimensions[get_column_letter(start_col)].width = 7  # C/A width
        sheet.column_dimensions[get_column_letter(start_col + 1)].width = (
            7  # S/T width
        )
        sheet.column_dimensions[get_column_letter(start_col + 2)].width = (
            7  # Exam width
        )
        sheet.column_dimensions[get_column_letter(start_col + 3)].width = (
            7  # Total width
        )

        sheet.merge_cells(
            start_row=2, start_column=start_col, end_row=2, end_column=end_col
        )
        cell = sheet.cell(row=2, column=start_col)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = sub_header_font

    # Add sub-headers (C/A, S/T, Exam, Total) under each student name
    sub_headers = ["Subjects"]
    for _ in broadsheet_data:
        sub_headers.extend(["C/A", "S/T", "Exam", "Total"])
    sub_headers.append("")
    sheet.append(sub_headers)

    # Write data for each subject and student
    for subject in subjects:
        row = [subject.name]
        for student_data in broadsheet_data:
            result = student_data["results"][subject.id]
            if result:
                row.extend(
                    [
                        result.class_assessment if result.class_assessment else "",
                        result.summative_test if result.summative_test else "",
                        result.exam if result.exam else "",
                        result.total if result.total else "",
                    ]
                )
            else:
                row.extend(["", "", "", ""])
        row.append(
            subject_averages[subject.id]["average"]
            if subject_averages[subject.id]["average"]
            else ""
        )
        sheet.append(row)

    # Write grand totals, averages, and positions, ensuring values are not empty
    sheet.append([""])
    sheet.append(
        ["Grand Total"]
        + sum(
            [
                [
                    "",
                    "",
                    "",
                    (
                        student_data["grand_total"]
                        if student_data["grand_total"]
                        else ""
                    ),
                ]
                for student_data in broadsheet_data
            ],
            [],
        )
        + [""]
    )
    sheet.append(
        ["Average"]
        + sum(
            [
                [
                    "",
                    "",
                    "",
                    student_data["average"] if student_data["average"] else "",
                ]
                for student_data in broadsheet_data
            ],
            [],
        )
        + [""]
    )
    sheet.append(
        ["Position"]
        + sum(
            [
                ["", "", "", student_data["position"]]
                for student_data in broadsheet_data
            ],
            [],
        )
        + [""]
    )

    # Set page orientation to landscape
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4

    # Margins (in inches)
    sheet.page_margins.top = 0.75  # Top margin: 1.91 cm
    sheet.page_margins.bottom = 0.75  # Bottom margin: 1.91 cm
    sheet.page_margins.left = 0.252  # Left margin: 0.64 cm
    sheet.page_margins.right = 0.252  # Right margin: 0.64 cm
    sheet.page_margins.header = 0.299  # Header margin: 0.76 cm
    sheet.page_margins.footer = 0.299  # Footer margin: 0.76 cm

    for row in sheet.iter_rows(
        min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column
    ):
        for cell in row:
            cell.border = border
            cell.font = cell_font

            # Apply left alignment specifically to "Subjects" column, sub-headers, and result cells
            if cell.column == 1:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = alignment

    # Save workbook to bytes buffer
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

